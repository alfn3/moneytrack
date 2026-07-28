import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

# ============================================================
# KONSTANTA WARNA & STYLE
# ============================================================
COLOR_HEADER_UNGU   = "FF4E5B91"   # Header utama ungu tua  (Sheet 1: DASHBOARD)
COLOR_HEADER_UNGU2  = "FF4E5B92"   # Header utama ungu tua  (Sheet 2,3,4,6)
COLOR_WHITE         = "FFFFFFFF"
COLOR_TEXT_DARK     = "FF1A1B20"
COLOR_BG_LIGHT      = "FFFAF8FF"   # Background putih keunguan muda (dashboard)
COLOR_BG_LAVENDER   = "FFEEECF3"   # Baris zebra lavender (dashboard kategori)
COLOR_BG_CARD       = "FFDEE1F9"   # Card header pemasukan/pengeluaran
COLOR_BG_SUBTITLE   = "FFE9E7EE"   # Subtitle / baris zebra sheet lain
COLOR_BG_PANDUAN    = "FFFAF7FF"   # Background panduan/ringkasan
COLOR_TRANSPARENT   = "00000000"   # Transparan / no fill

FORMAT_RP       = '[$Rp ]#,##0'
FORMAT_RP_RED   = '"Rp"#,##0;[RED]"(Rp"#,##0\\);\\-'
FORMAT_PCT      = '0.0%'

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_font(color=COLOR_TEXT_DARK, bold=False, size=11, name="Calibri"):
    return Font(name=name, color=color, bold=bold, size=size)

def make_align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, coord, value, fill_hex=None, font=None, align=None, num_format=None):
    c = ws[coord]
    c.value = value
    if fill_hex:
        c.fill = make_fill(fill_hex)
    if font:
        c.font = font
    if align:
        c.alignment = align
    if num_format:
        c.number_format = num_format
    return c

# ============================================================
# BUAT WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: DASHBOARD LAPORAN
# ============================================================
ws1 = wb.active
ws1.title = "DASHBOARD LAPORAN"

# --- Lebar Kolom ---
col_widths_dash = {
    'A': 17.29, 'B': 13.0, 'C': 13.0, 'D': 13.0,
    'E': 13.0,  'F': 13.0, 'G': 13.0, 'H': 13.0, 'I': 8.71,
}
for col, w in col_widths_dash.items():
    ws1.column_dimensions[col].width = w
for col in [get_column_letter(i) for i in range(10, 27)]:
    ws1.column_dimensions[col].width = 13.0

# --- Tinggi Baris ---
ws1.row_dimensions[1].height = 27.0
ws1.row_dimensions[6].height = 45.0
for r in range(21, 1001):
    ws1.row_dimensions[r].height = 15.75

# --- ROW 1: Judul ---
ws1.merge_cells('A1:H1')
set_cell(ws1, 'A1', "DASHBOARD LAPORAN KEUANGAN PRIBADI",
         fill_hex=COLOR_HEADER_UNGU,
         font=make_font(color=COLOR_WHITE, bold=True, size=14),
         align=make_align("center", "center"))

# --- ROW 3: Pilih Bulan ---
set_cell(ws1, 'A3', "Pilih Bulan:",
         fill_hex=COLOR_BG_LIGHT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
set_cell(ws1, 'B3', "Jan",
         fill_hex=COLOR_BG_LIGHT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align("center", "bottom"),
         num_format=FORMAT_RP)
# C3 adalah ArrayFormula
ws1['C3'] = ArrayFormula('C3', "=INDEX(SETTINGS!$J$4:$J$15,MATCH($B$3,SETTINGS!$I$4:$I$15,0))")
ws1['C3'].fill = make_fill(COLOR_BG_LIGHT)
ws1['C3'].font = make_font(color=COLOR_TEXT_DARK, size=11)
ws1['C3'].alignment = make_align(h=None, v="bottom")

# --- ROW 5: Card Labels ---
ws1.merge_cells('A5:B5')
ws1.merge_cells('C5:D5')
ws1.merge_cells('E5:F5')
ws1.merge_cells('G5:H5')
for coord, label in [('A5','PEMASUKAN'), ('C5','PENGELUARAN'), ('E5','TABUNGAN'), ('G5','SISA BUDGET')]:
    set_cell(ws1, coord, label,
             fill_hex=COLOR_BG_CARD,
             font=make_font(color=COLOR_TEXT_DARK, bold=True),
             align=make_align("center","center"))

# --- ROW 6: Card Values ---
ws1.merge_cells('A6:B6')
ws1.merge_cells('C6:D6')
ws1.merge_cells('E6:F6')
ws1.merge_cells('G6:H6')

# A6: Pemasukan (SUMIFS)
ws1['A6'] = "=SUMIFS('CATATAN PENGELUARAN'!$F$4:$F$153,'CATATAN PENGELUARAN'!$H$4:$H$153,$C$3,'CATATAN PENGELUARAN'!$B$4:$B$153,\"Pemasukan\")"
ws1['A6'].fill = make_fill(COLOR_BG_LIGHT)
ws1['A6'].font = make_font(color=COLOR_TEXT_DARK, bold=True, size=18)
ws1['A6'].alignment = make_align("center","center")
ws1['A6'].number_format = FORMAT_RP

# C6: Pengeluaran (SUMIFS)
ws1['C6'] = "=SUMIFS('CATATAN PENGELUARAN'!$F$4:$F$153,'CATATAN PENGELUARAN'!$H$4:$H$153,$C$3,'CATATAN PENGELUARAN'!$B$4:$B$153,\"Pengeluaran\")"
ws1['C6'].fill = make_fill(COLOR_BG_LIGHT)
ws1['C6'].font = make_font(color=COLOR_TEXT_DARK, bold=True, size=18)
ws1['C6'].alignment = make_align("center","center")
ws1['C6'].number_format = FORMAT_RP

# E6: Tabungan = A6 - C6
ws1['E6'] = "=A6-C6"
ws1['E6'].fill = make_fill(COLOR_BG_LIGHT)
ws1['E6'].font = make_font(color=COLOR_TEXT_DARK, bold=True, size=18)
ws1['E6'].alignment = make_align("center","center")
ws1['E6'].number_format = FORMAT_RP

# G6: Sisa Budget = ArrayFormula (budget bulan ini - pengeluaran)
ws1['G6'] = ArrayFormula('G6', "=IFERROR(INDEX('BUDGETING TAHUNAN'!$B$15:$M$15,1,MATCH($B$3,'BUDGETING TAHUNAN'!$B$3:$M$3,0)),0)-C6")
ws1['G6'].fill = make_fill(COLOR_BG_LIGHT)
ws1['G6'].font = make_font(color=COLOR_TEXT_DARK, bold=True, size=18)
ws1['G6'].alignment = make_align("center","center")
ws1['G6'].number_format = FORMAT_RP

# --- ROW 9: Section Headers ---
ws1.merge_cells('A9:D9')
ws1.merge_cells('F9:H9')
set_cell(ws1, 'A9', "PENGELUARAN PER KATEGORI (BULAN TERPILIH)",
         fill_hex=COLOR_HEADER_UNGU,
         font=make_font(color=COLOR_WHITE, bold=True),
         align=make_align("center","center"))
set_cell(ws1, 'F9', "TREND BULANAN (SETAHUN)",
         fill_hex=COLOR_HEADER_UNGU,
         font=make_font(color=COLOR_WHITE, bold=True),
         align=make_align("center","center"))

# --- ROW 10: Column Headers ---
ws1.merge_cells('A10:B10')
ws1.merge_cells('C10:D10')
for coord, label in [('A10','Kategori'), ('C10','Realisasi'), ('F10','Bulan'), ('G10','Pemasukan'), ('H10','Pengeluaran')]:
    set_cell(ws1, coord, label,
             fill_hex=COLOR_HEADER_UNGU,
             font=make_font(color=COLOR_WHITE, bold=True),
             align=make_align("center","center"))

# --- ROW 11-21: Kategori data ---
kategori_list = [
    "Kebutuhan Pokok", "Makanan & Minuman", "Transportasi", "Kesehatan",
    "Hiburan", "Pendidikan", "Utilitas & Listrik", "Cicilan",
    "Belanja Pribadi", "Sosial & Amal", "Lainnya"
]
ringkasan_rows = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]  # Baris di RINGKASAN BUDGETING
bulan_list = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]

for i, (kat, rb) in enumerate(zip(kategori_list, ringkasan_rows)):
    row = 11 + i
    bg = COLOR_BG_LAVENDER if i % 2 == 0 else COLOR_WHITE

    ws1.merge_cells(f'A{row}:B{row}')
    ws1.merge_cells(f'C{row}:D{row}')

    set_cell(ws1, f'A{row}', kat,
             fill_hex=bg,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align(h=None, v="bottom"))
    set_cell(ws1, f'C{row}', f"='RINGKASAN BUDGETING'!C{rb}",
             fill_hex=bg,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_RP)

# Trend bulanan (kolom F-H, baris 11-22)
for j, bln in enumerate(bulan_list):
    row = 11 + j
    bulan_num = j + 1
    set_cell(ws1, f'F{row}', bln,
             fill_hex=COLOR_BG_LIGHT,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align("center","bottom"))

    f_pem = (f"=SUMIFS('CATATAN PENGELUARAN'!$F$4:$F$153,"
             f"'CATATAN PENGELUARAN'!$H$4:$H$153,{bulan_num},"
             f"'CATATAN PENGELUARAN'!$B$4:$B$153,\"Pemasukan\")")
    set_cell(ws1, f'G{row}', f_pem,
             fill_hex=COLOR_BG_LIGHT,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_RP)

    f_peng = (f"=SUMIFS('CATATAN PENGELUARAN'!$F$4:$F$153,"
              f"'CATATAN PENGELUARAN'!$H$4:$H$153,{bulan_num},"
              f"'CATATAN PENGELUARAN'!$B$4:$B$153,\"Pengeluaran\")")
    set_cell(ws1, f'H{row}', f_peng,
             fill_hex=COLOR_BG_LIGHT,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_RP)

# ============================================================
# SHEET 2: BUDGETING TAHUNAN
# ============================================================
ws2 = wb.create_sheet("BUDGETING TAHUNAN")

# --- Lebar Kolom ---
ws2.column_dimensions['A'].width = 31.0
ws2.column_dimensions['B'].width = 18.14
for col in [get_column_letter(i) for i in range(3, 14)]:  # C-M
    ws2.column_dimensions[col].width = 13.0
ws2.column_dimensions['N'].width = 8.71
for col in [get_column_letter(i) for i in range(15, 27)]:
    ws2.column_dimensions[col].width = 13.0

# --- Tinggi Baris ---
for r in range(21, 1001):
    ws2.row_dimensions[r].height = 15.75

# --- ROW 1: Judul ---
ws2.merge_cells('A1:M1')
set_cell(ws2, 'A1', "RENCANA BUDGETING TAHUNAN",
         fill_hex=COLOR_HEADER_UNGU2,
         font=make_font(color=COLOR_WHITE, bold=True),
         align=make_align("center","bottom"))

# --- ROW 2: Subtitle ---
ws2.merge_cells('A2:M2')
set_cell(ws2, 'A2', "Isi rencana anggaran per kategori untuk tiap bulan. Angka ini digunakan sebagai patokan budget di sheet Ringkasan.",
         fill_hex=COLOR_BG_SUBTITLE,
         font=make_font(color=COLOR_TEXT_DARK, size=9),
         align=make_align(h=None, v="bottom"))

# --- ROW 3: Header bulan ---
bulan_headers = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
set_cell(ws2, 'A3', "Kategori",
         fill_hex=COLOR_WHITE,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=11),
         align=make_align("center","center"))
for col_idx, bln in enumerate(bulan_headers, 2):
    col_letter = get_column_letter(col_idx)
    set_cell(ws2, f'{col_letter}3', bln,
             fill_hex=COLOR_WHITE,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=11),
             align=make_align("center","center"),
             num_format=FORMAT_RP)

# --- ROW 4-14: Kategori pengeluaran ---
kategori_budget = [
    "Kebutuhan Pokok", "Makanan & Minuman", "Transportasi", "Kesehatan",
    "Hiburan", "Pendidikan", "Utilitas & Listrik", "Cicilan",
    "Belanja Pribadi", "Sosial & Amal", "Lainnya"
]
for i, kat in enumerate(kategori_budget):
    row = 4 + i
    bg = COLOR_BG_SUBTITLE if i % 2 == 0 else COLOR_WHITE
    set_cell(ws2, f'A{row}', kat,
             fill_hex=bg,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align(h=None, v="bottom"))
    for col_idx in range(2, 14):
        col_letter = get_column_letter(col_idx)
        set_cell(ws2, f'{col_letter}{row}', 0.0,
                 fill_hex=bg,
                 font=make_font(color=COLOR_TEXT_DARK, size=10),
                 align=make_align("center","bottom"),
                 num_format=FORMAT_RP)

# --- ROW 15: TOTAL PENGELUARAN ---
set_cell(ws2, 'A15', "TOTAL PENGELUARAN",
         fill_hex=COLOR_WHITE,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
for col_idx in range(2, 14):
    col_letter = get_column_letter(col_idx)
    set_cell(ws2, f'{col_letter}15', f'=SUM({col_letter}4:{col_letter}14)',
             fill_hex=COLOR_WHITE,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_RP)

# --- ROW 16: TARGET TABUNGAN ---
set_cell(ws2, 'A16', "TARGET TABUNGAN",
         fill_hex=COLOR_BG_SUBTITLE,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
for col_idx in range(2, 14):
    col_letter = get_column_letter(col_idx)
    set_cell(ws2, f'{col_letter}16', 0.0,
             fill_hex=COLOR_BG_SUBTITLE,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_RP)

# --- ROW 17: RENCANA PEMASUKAN ---
set_cell(ws2, 'A17', "RENCANA PEMASUKAN",
         fill_hex=COLOR_WHITE,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
for col_idx in range(2, 14):
    col_letter = get_column_letter(col_idx)
    set_cell(ws2, f'{col_letter}17', 0.0,
             fill_hex=COLOR_WHITE,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_RP)

# ============================================================
# SHEET 3: CATATAN PENGELUARAN
# ============================================================
ws3 = wb.create_sheet("CATATAN PENGELUARAN")

# --- Lebar Kolom ---
col_widths_tx = {'A': 17.14, 'B': 14.86, 'C': 14.0, 'D': 16.14,
                 'E': 11.71, 'F': 11.14, 'G': 21.43, 'H': 8.0, 'I': 8.71}
for col, w in col_widths_tx.items():
    ws3.column_dimensions[col].width = w
for col in [get_column_letter(i) for i in range(10, 27)]:
    ws3.column_dimensions[col].width = 13.0

# --- Tinggi Baris ---
for r in range(21, 1001):
    ws3.row_dimensions[r].height = 15.75

# --- ROW 1: Judul ---
ws3.merge_cells('A1:H1')
set_cell(ws3, 'A1', "CATATAN TRANSAKSI (PEMASUKAN & PENGELUARAN)",
         fill_hex=COLOR_HEADER_UNGU2,
         font=make_font(color=COLOR_WHITE, bold=True),
         align=make_align("center","bottom"))

# --- ROW 2: Subtitle ---
ws3.merge_cells('A2:H2')
set_cell(ws3, 'A2', "Kolom H (Bulan) otomatis, jangan diedit. Kategori harus sesuai daftar di SETTINGS. Isi dari baris 4.",
         fill_hex="FFFAF7FF",
         font=make_font(color=COLOR_TEXT_DARK, size=9),
         align=make_align(h=None, v="bottom"))

# --- ROW 3: Header kolom ---
headers_tx = ["Tanggal", "Tipe", "Kategori", "Keterangan", "Akun", "Jumlah", "Saldo Berjalan", "Bulan"]
for col_idx, hdr in enumerate(headers_tx, 1):
    col_letter = get_column_letter(col_idx)
    set_cell(ws3, f'{col_letter}3', hdr,
             fill_hex=COLOR_HEADER_UNGU2,
             font=make_font(color=COLOR_WHITE, bold=True),
             align=make_align("center","center"))

# --- ROW 4-153: Formula Saldo Berjalan & Bulan ---
for row in range(4, 154):
    prev = row - 1
    # G: Saldo Berjalan
    if row == 4:
        saldo_formula = f'=IF(A{row}="","",N(G3)+IF(B{row}="Pemasukan",F{row},-F{row}))'
    else:
        saldo_formula = f'=IF(A{row}="","",N(G{prev})+IF(B{row}="Pemasukan",F{row},-F{row}))'

    ws3[f'G{row}'] = saldo_formula
    ws3[f'G{row}'].fill = make_fill(COLOR_TRANSPARENT)
    ws3[f'G{row}'].font = make_font(color=COLOR_TEXT_DARK, size=11)
    ws3[f'G{row}'].alignment = make_align(h=None, v="bottom")
    ws3[f'G{row}'].number_format = FORMAT_RP_RED

    # H: Bulan
    ws3[f'H{row}'] = f'=IF(A{row}="","",MONTH(A{row}))'
    ws3[f'H{row}'].fill = make_fill(COLOR_TRANSPARENT)
    ws3[f'H{row}'].font = make_font(color=COLOR_TEXT_DARK, size=11)
    ws3[f'H{row}'].alignment = make_align(h=None, v="bottom")

# ============================================================
# SHEET 4: RINGKASAN BUDGETING
# ============================================================
ws4 = wb.create_sheet("RINGKASAN BUDGETING")

# --- Lebar Kolom ---
ws4.column_dimensions['A'].width = 38.57
ws4.column_dimensions['B'].width = 15.43
ws4.column_dimensions['C'].width = 13.0
ws4.column_dimensions['D'].width = 13.0
ws4.column_dimensions['E'].width = 15.57
ws4.column_dimensions['F'].width = 8.71
for col in [get_column_letter(i) for i in range(7, 27)]:
    ws4.column_dimensions[col].width = 13.0

# --- Tinggi Baris ---
for r in range(21, 1001):
    ws4.row_dimensions[r].height = 15.75

# --- ROW 1: Judul ---
ws4.merge_cells('A1:E1')
set_cell(ws4, 'A1', "RINGKASAN: BUDGET vs REALISASI",
         fill_hex=COLOR_BG_PANDUAN,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=12),
         align=make_align(h=None, v="bottom"))

# --- ROW 3: Pilih Bulan ---
set_cell(ws4, 'A3', "Pilih Bulan:",
         fill_hex=COLOR_BG_PANDUAN,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
set_cell(ws4, 'B3', "Jan",
         fill_hex=COLOR_BG_PANDUAN,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align("center","bottom"),
         num_format=FORMAT_RP)

# --- ROW 5: Header tabel ---
for coord, label in [('A5','Kategori'), ('B5','Budget'), ('C5','Realisasi'), ('D5','Sisa'), ('E5','Persentase')]:
    set_cell(ws4, coord, label,
             fill_hex=COLOR_HEADER_UNGU2,
             font=make_font(color=COLOR_WHITE, bold=True),
             align=make_align("center","center"),
             num_format=FORMAT_RP if label in ('Budget','Realisasi') else 'General')

# --- ROW 6-16: Kategori dengan formula ---
for i, kat in enumerate(kategori_budget):
    row = 6 + i

    set_cell(ws4, f'A{row}', kat,
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align(h=None, v="bottom"))

    # B: Budget (ArrayFormula - INDEX/MATCH dari BUDGETING TAHUNAN)
    b_formula = (f"=IFERROR(INDEX('BUDGETING TAHUNAN'!$B$4:$M$14,"
                 f"MATCH($A{row},'BUDGETING TAHUNAN'!$A$4:$A$14,0),"
                 f"MATCH($B$3,'BUDGETING TAHUNAN'!$B$3:$M$3,0)),0)")
    ws4[f'B{row}'] = ArrayFormula(f'B{row}', b_formula)
    ws4[f'B{row}'].fill = make_fill(COLOR_TRANSPARENT)
    ws4[f'B{row}'].font = make_font(color=COLOR_TEXT_DARK, size=10)
    ws4[f'B{row}'].alignment = make_align("center","bottom")
    ws4[f'B{row}'].number_format = FORMAT_RP

    # C: Realisasi (ArrayFormula - SUMIFS dari CATATAN PENGELUARAN)
    c_formula = (f"=SUMIFS('CATATAN PENGELUARAN'!$F$4:$F$153,"
                 f"'CATATAN PENGELUARAN'!$C$4:$C$153,$A{row},"
                 f"'CATATAN PENGELUARAN'!$H$4:$H$153,"
                 f"INDEX(SETTINGS!$J$4:$J$15,MATCH($B$3,SETTINGS!$I$4:$I$15,0)),"
                 f"'CATATAN PENGELUARAN'!$B$4:$B$153,\"Pengeluaran\")")
    ws4[f'C{row}'] = ArrayFormula(f'C{row}', c_formula)
    ws4[f'C{row}'].fill = make_fill(COLOR_TRANSPARENT)
    ws4[f'C{row}'].font = make_font(color=COLOR_TEXT_DARK, size=10)
    ws4[f'C{row}'].alignment = make_align("center","bottom")
    ws4[f'C{row}'].number_format = FORMAT_RP

    # D: Sisa = Budget - Realisasi
    set_cell(ws4, f'D{row}', f'=B{row}-C{row}',
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_RP)

    # E: Persentase
    set_cell(ws4, f'E{row}', f'=IFERROR(C{row}/B{row},0)',
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align("center","bottom"),
             num_format=FORMAT_PCT)

# --- ROW 17: TOTAL ---
set_cell(ws4, 'A17', "TOTAL",
         fill_hex=COLOR_TRANSPARENT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
set_cell(ws4, 'B17', "=SUM(B6:B16)",
         fill_hex=COLOR_TRANSPARENT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align("center","bottom"),
         num_format=FORMAT_RP)
set_cell(ws4, 'C17', "=SUM(C6:C16)",
         fill_hex=COLOR_TRANSPARENT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align("center","bottom"),
         num_format=FORMAT_RP)
set_cell(ws4, 'D17', "=B17-C17",
         fill_hex=COLOR_TRANSPARENT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align("center","bottom"),
         num_format=FORMAT_RP)
set_cell(ws4, 'E17', "=IFERROR(C17/B17,0)",
         fill_hex=COLOR_TRANSPARENT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align("center","bottom"),
         num_format=FORMAT_PCT)

# ============================================================
# SHEET 5: TRACKER ASET
# ============================================================
ws5 = wb.create_sheet("TRACKER ASET")

# --- Lebar Kolom ---
col_widths_aset = {'A': 30.29, 'B': 17.57, 'C': 18.14, 'D': 9.71,
                   'E': 16.29, 'F': 19.57, 'G': 8.71}
for col, w in col_widths_aset.items():
    ws5.column_dimensions[col].width = w
for col in [get_column_letter(i) for i in range(8, 27)]:
    ws5.column_dimensions[col].width = 13.0

# --- Tinggi Baris ---
for r in range(21, 1001):
    ws5.row_dimensions[r].height = 15.75

# --- ROW 1: Judul ---
ws5.merge_cells('A1:F1')
set_cell(ws5, 'A1', "PENCATATAN ASET",
         fill_hex=COLOR_BG_PANDUAN,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=12),
         align=make_align(h=None, v="bottom"))

# --- ROW 3: Header ASET LIKUID ---
ws5.merge_cells('A3:F3')
set_cell(ws5, 'A3', "ASET LIKUID (Mudah Dicairkan)",
         fill_hex=COLOR_HEADER_UNGU2,
         font=make_font(color=COLOR_WHITE, bold=True),
         align=make_align("center","bottom"))

# --- ROW 4: Sub-header LIKUID ---
for coord, label, fmt in [
    ('A4','Jenis Aset','General'), ('B4','Nilai Saat Ini',FORMAT_RP),
    ('C4','Target',FORMAT_RP), ('D4','Progress','General'),
    ('E4','Tanggal Update','General'), ('F4','Catatan','General')
]:
    set_cell(ws5, coord, label,
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=11),
             align=make_align("center","center"),
             num_format=fmt)

# --- ROW 5-7: Aset Likuid ---
aset_likuid = ["Tabungan Bank", "E-Wallet", "Uang Tunai"]
for i, aset in enumerate(aset_likuid):
    row = 5 + i
    set_cell(ws5, f'A{row}', aset,
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, size=11),
             align=make_align(h=None, v="bottom"))
    for col in ['B', 'C']:
        set_cell(ws5, f'{col}{row}', 0.0,
                 fill_hex=COLOR_TRANSPARENT,
                 font=make_font(color=COLOR_TEXT_DARK, size=11),
                 align=make_align("center","bottom"),
                 num_format=FORMAT_RP)
    set_cell(ws5, f'D{row}', f'=IFERROR(B{row}/C{row},0)',
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, size=11),
             align=make_align("center","bottom"),
             num_format=FORMAT_PCT)

# --- ROW 8: Spacer (kosong) ---

# --- ROW 9: Header ASET NON-LIKUID ---
ws5.merge_cells('A9:F9')
set_cell(ws5, 'A9', "ASET NON-LIKUID (Investasi)",
         fill_hex=COLOR_HEADER_UNGU2,
         font=make_font(color=COLOR_WHITE, bold=True),
         align=make_align("center","bottom"))

# --- ROW 10: Sub-header NON-LIKUID ---
for coord, label, fmt in [
    ('A10','Jenis Aset','General'), ('B10','Nilai Saat Ini',FORMAT_RP),
    ('C10','Target',FORMAT_RP), ('D10','Progress','General'),
    ('E10','Tanggal Update','General'), ('F10','Catatan','General')
]:
    set_cell(ws5, coord, label,
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=11),
             align=make_align("center","center"),
             num_format=fmt)

# --- ROW 11-15: Aset Non-Likuid ---
aset_non_likuid = ["Emas", "Saham", "Crypto", "Deposito", "Reksa Dana"]
for i, aset in enumerate(aset_non_likuid):
    row = 11 + i
    set_cell(ws5, f'A{row}', aset,
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, size=11),
             align=make_align(h=None, v="bottom"))
    for col in ['B', 'C']:
        set_cell(ws5, f'{col}{row}', 0.0,
                 fill_hex=COLOR_TRANSPARENT,
                 font=make_font(color=COLOR_TEXT_DARK, size=11),
                 align=make_align("center","bottom"),
                 num_format=FORMAT_RP)
    set_cell(ws5, f'D{row}', f'=IFERROR(B{row}/C{row},0)',
             fill_hex=COLOR_TRANSPARENT,
             font=make_font(color=COLOR_TEXT_DARK, size=11),
             align=make_align("center","bottom"),
             num_format=FORMAT_PCT)

# --- ROW 16: Spacer ---

# --- ROW 17: TOTAL ASET ---
# Named ranges untuk Aset_Likuid (B5:B7) dan Aset_NonLikuid (B11:B15)
set_cell(ws5, 'A17', "TOTAL ASET KESELURUHAN",
         fill_hex=COLOR_TRANSPARENT,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
ws5['B17'] = "=SUM(B5:B7)+SUM(B11:B15)"
ws5['B17'].fill = make_fill(COLOR_TRANSPARENT)
ws5['B17'].font = make_font(color=COLOR_TEXT_DARK, bold=True, size=10)
ws5['B17'].alignment = make_align("center","bottom")
ws5['B17'].number_format = FORMAT_RP

# ============================================================
# SHEET 6: SETTINGS
# ============================================================
ws6 = wb.create_sheet("SETTINGS")

# --- Lebar Kolom ---
col_widths_set = {'A': 65.86, 'B': 2.0, 'C': 20.71, 'D': 2.0,
                  'E': 17.29, 'F': 2.0, 'G': 16.0, 'H': 2.0,
                  'I': 7.29, 'J': 3.86, 'K': 8.71}
for col, w in col_widths_set.items():
    ws6.column_dimensions[col].width = w
for col in [get_column_letter(i) for i in range(12, 27)]:
    ws6.column_dimensions[col].width = 13.0

# --- Tinggi Baris ---
for r in range(21, 1001):
    ws6.row_dimensions[r].height = 15.75

# --- ROW 1: Judul ---
ws6.merge_cells('A1:J1')
set_cell(ws6, 'A1', "PENGATURAN & KONFIGURASI",
         fill_hex=COLOR_HEADER_UNGU2,
         font=make_font(color=COLOR_WHITE, bold=True),
         align=make_align("center","bottom"))

# --- ROW 2: Subtitle ---
ws6.merge_cells('A2:J2')
set_cell(ws6, 'A2', "Ubah daftar di bawah ini untuk menyesuaikan kategori, sumber pemasukan, dan rekening yang digunakan.",
         fill_hex=COLOR_BG_SUBTITLE,
         font=make_font(color=COLOR_TEXT_DARK, size=9),
         align=make_align(h=None, v="bottom"))

# --- ROW 3: Sub-headers ---
for coord, label in [('A3','KATEGORI PENGELUARAN'), ('C3','SUMBER PEMASUKAN'),
                      ('E3','REKENING BANK'), ('G3','TIPE TRANSAKSI'),
                      ('I3','BULAN'), ('J3','NO')]:
    set_cell(ws6, coord, label,
             fill_hex=COLOR_WHITE,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
             align=make_align(h=None, v="bottom"))

# --- ROW 4-15: Data settings ---
kategori_set = ["Kebutuhan Pokok", "Makanan & Minuman", "Transportasi", "Kesehatan",
                "Hiburan", "Pendidikan", "Utilitas & Listrik", "Cicilan",
                "Belanja Pribadi", "Sosial & Amal", "Lainnya", ""]
sumber = ["Gaji Utama", "Gaji Tambahan", "Bonus", "Investasi Return", "Lainnya", "",
          "", "", "", "", "", ""]
rekening = ["Rekening Utama", "Rekening Tabungan", "E-Wallet", "Tunai",
            "Rekening 5", "Rekening 6", "", "", "", "", "", ""]
tipe = ["Pemasukan", "Pengeluaran", "", "", "", "", "", "", "", "", "", ""]
bulan_set = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]

for i in range(12):
    row = 4 + i
    bg = COLOR_BG_SUBTITLE if i % 2 == 0 else COLOR_WHITE

    if kategori_set[i]:
        set_cell(ws6, f'A{row}', kategori_set[i],
                 fill_hex=bg,
                 font=make_font(color=COLOR_TEXT_DARK, size=10),
                 align=make_align(h=None, v="bottom"))
    if sumber[i]:
        set_cell(ws6, f'C{row}', sumber[i],
                 fill_hex=bg,
                 font=make_font(color=COLOR_TEXT_DARK, size=10),
                 align=make_align(h=None, v="bottom"))
    if rekening[i]:
        set_cell(ws6, f'E{row}', rekening[i],
                 fill_hex=bg,
                 font=make_font(color=COLOR_TEXT_DARK, size=10),
                 align=make_align(h=None, v="bottom"))
    if tipe[i]:
        set_cell(ws6, f'G{row}', tipe[i],
                 fill_hex=bg,
                 font=make_font(color=COLOR_TEXT_DARK, size=10),
                 align=make_align(h=None, v="bottom"))

    set_cell(ws6, f'I{row}', bulan_set[i],
             fill_hex=bg,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align(h=None, v="bottom"))
    set_cell(ws6, f'J{row}', float(i + 1),
             fill_hex=bg,
             font=make_font(color=COLOR_TEXT_DARK, size=10),
             align=make_align(h=None, v="bottom"))

# ============================================================
# SHEET 7: PANDUAN
# ============================================================
ws7 = wb.create_sheet("PANDUAN")

# --- Lebar Kolom ---
ws7.column_dimensions['A'].width = 95.86
ws7.column_dimensions['B'].width = 8.71
for col in [get_column_letter(i) for i in range(3, 27)]:
    ws7.column_dimensions[col].width = 13.0

# --- Tinggi Baris ---
for r in [4, 7, 10, 13, 16, 19, 22]:
    ws7.row_dimensions[r].height = 45.0
for r in range(21, 1001):
    ws7.row_dimensions[r].height = 15.75

# --- ROW 1: Judul ---
ws7.merge_cells('A1:D1')
set_cell(ws7, 'A1', "PANDUAN PENGGUNAAN TEMPLATE BUDGETING",
         fill_hex=COLOR_BG_PANDUAN,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=12),
         align=make_align(h=None, v="bottom"))

panduan_items = [
    ("1. SETTINGS",
     "Atur kategori, sumber pemasukan, rekening bank, dan tipe transaksi yang kamu gunakan. Nama kategori HARUS sama persis dengan yang dipakai di Catatan Pengeluaran."),
    ("2. BUDGETING TAHUNAN",
     "Isi rencana anggaran per kategori untuk tiap bulan. Angka ini digunakan sebagai patokan budget di sheet Ringkasan."),
    ("3. CATATAN PENGELUARAN",
     "Catat SETIAP transaksi (pemasukan maupun pengeluaran). Kolom Bulan (H) akan terisi otomatis. Jangan hapus atau pindahkan kolom H."),
    ("4. RINGKASAN BUDGETING",
     "Pilih bulan di sel B3, lalu tabel otomatis membandingkan budget (dari Budgeting Tahunan) vs realisasi (dari Catatan Pengeluaran)."),
    ("5. TRACKER ASET",
     "Catat aset likuid (tabungan, e-wallet, tunai) dan non-likuid (emas, saham, crypto, dll). Update secara berkala untuk memantau perkembangan kekayaan."),
    ("6. DASHBOARD LAPORAN",
     "Pilih bulan di sel B3 untuk melihat ringkasan Pemasukan, Pengeluaran, Tabungan, dan Sisa Budget bulan tersebut, beserta trend tahunan."),
]

for i, (header, desc) in enumerate(panduan_items):
    base_row = 3 + (i * 3)
    set_cell(ws7, f'A{base_row}', header,
             fill_hex=COLOR_BG_PANDUAN,
             font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
             align=make_align(h=None, v="bottom"))
    ws7.merge_cells(f'A{base_row+1}:D{base_row+1}')
    set_cell(ws7, f'A{base_row+1}', desc,
             fill_hex=COLOR_BG_PANDUAN,
             font=make_font(color=COLOR_TEXT_DARK, size=11),
             align=make_align(h=None, v="bottom", wrap=True))

# --- CATATAN PENTING ---
set_cell(ws7, 'A21', "CATATAN PENTING",
         fill_hex=COLOR_BG_PANDUAN,
         font=make_font(color=COLOR_TEXT_DARK, bold=True, size=10),
         align=make_align(h=None, v="bottom"))
ws7.merge_cells('A22:D22')
set_cell(ws7, 'A22',
         "Jangan hapus/pindah kolom H (Bulan) di Catatan Pengeluaran. Kolom ini digunakan oleh semua formula di Dashboard dan Ringkasan untuk memfilter data per bulan.",
         fill_hex=COLOR_BG_PANDUAN,
         font=make_font(color=COLOR_TEXT_DARK, size=11),
         align=make_align(h=None, v="bottom", wrap=True))

# ============================================================
# SIMPAN FILE
# ============================================================
wb.save("template_budget.xlsx")
print("✅ File berhasil disimpan: template_budget.xlsx")
print(f"   Total sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"   - {s}")